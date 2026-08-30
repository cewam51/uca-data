import csv
from datetime import date, datetime, time
from itertools import islice
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class WorkbookTableError(ValueError):
    pass


def convert_xlsx_to_csv(source: Path, destination: Path) -> dict[str, Any]:
    """Extrait en flux la feuille la plus tabulaire d’un classeur XLSX."""
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError) as error:
        raise WorkbookTableError("Le classeur Excel n’est pas lisible.") from error

    try:
        candidate = _select_sheet(workbook)
        if candidate is None:
            raise WorkbookTableError("Le classeur Excel ne contient aucune feuille tabulaire exploitable.")
        sheet, header_index, width = candidate
        written_rows = 0
        with destination.open("x", encoding="utf-8", newline="") as output:
            writer = csv.writer(output)
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_index < header_index:
                    continue
                values = list(row[:width])
                values.extend([None] * (width - len(values)))
                if row_index == header_index:
                    writer.writerow(_unique_headers(values))
                    continue
                if not any(value is not None and str(value).strip() for value in values):
                    continue
                writer.writerow([_serializable(value) for value in values])
                written_rows += 1
        if written_rows == 0:
            destination.unlink(missing_ok=True)
            raise WorkbookTableError("La feuille Excel retenue ne contient aucune ligne de données.")
        return {"sheet_name": sheet.title, "converted_row_count": written_rows}
    finally:
        workbook.close()


def _select_sheet(workbook) -> tuple[Any, int, int] | None:
    candidates: list[tuple[tuple[int, int, int], int, Any, int, int]] = []
    for sheet_index, sheet in enumerate(workbook.worksheets):
        if sheet.sheet_state != "visible":
            continue
        sample = list(islice(sheet.iter_rows(values_only=True), 500))
        populated = [
            (row_index, row, _last_populated_column(row))
            for row_index, row in enumerate(sample)
            if _last_populated_column(row) > 0
        ]
        if len(populated) < 2:
            continue
        header_candidates = populated[:50]
        maximum_values = max(
            sum(_has_value(value) for value in row)
            for _, row, _ in header_candidates
        )
        minimum_header_values = max(1, (maximum_values * 3 + 4) // 5)
        header_index, _, header_width = next(
            item
            for item in header_candidates
            if sum(_has_value(value) for value in item[1]) >= minimum_header_values
        )
        rows_after_header = sum(1 for row_index, _, _ in populated if row_index > header_index)
        width = max(header_width, *(last_column for _, _, last_column in populated))
        if rows_after_header == 0 or width == 0:
            continue
        score = (rows_after_header, width, len(populated))
        candidates.append((score, -sheet_index, sheet, header_index, width))
    if not candidates:
        return None
    _, _, sheet, header_index, width = max(candidates, key=lambda item: (item[0], item[1]))
    return sheet, header_index, width


def _last_populated_column(row: tuple[Any, ...]) -> int:
    for index in range(len(row) - 1, -1, -1):
        if _has_value(row[index]):
            return index + 1
    return 0


def _has_value(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def _unique_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    occurrences: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = str(value).strip() if _has_value(value) else f"Colonne {index}"
        key = base.casefold()
        occurrences[key] = occurrences.get(key, 0) + 1
        headers.append(base if occurrences[key] == 1 else f"{base} ({occurrences[key]})")
    return headers


def _serializable(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)
